import openpyxl
import json
import os
import re

class OptimizedRobotConfigProcessor:
    def __init__(self):
        self.json_template = None
        self.load_json_template()
    
    def load_json_template(self):
        """加载JSON模板"""
        try:
            with open('d:/received/RA-IC_I-A-1A3BH0.json', 'r', encoding='utf-8') as f:
                self.json_template = json.load(f)
            print("JSON模板加载成功")
        except Exception as e:
            print(f"加载JSON模板失败: {e}")
            self.json_template = None
    
    def extract_board_model(self, c_column_value):
        """从C列内容中提取单板型号"""
        if not c_column_value:
            return None
            
        c_text = str(c_column_value)
        
        # 查找"单板型号："格式
        model_match = re.search(r'单板型号：([^\n]+)', c_text)
        if model_match:
            return model_match.group(1).strip()
        
        # 查找"型号："格式
        model_match2 = re.search(r'型号：([^\n]+)', c_text)
        if model_match2:
            return model_match2.group(1).strip()
            
        # 如果没有找到特定格式，返回整个文本的前几部分
        lines = c_text.split('\n')
        if lines:
            return lines[0].strip()
            
        return None
    
    def analyze_excel_for_boards(self):
        """分析Excel文件，识别所有单板"""
        print("分析Excel文件结构...")
        
        wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
        ws = wb['整机配置表']
        
        boards = []
        current_board = None
        board_start_row = None
        
        # 查找所有单板开始行（C列包含"单板型号"或"型号："的行）
        for i in range(1, 200):
            try:
                a_val = ws.cell(row=i, column=1).value
                c_val = ws.cell(row=i, column=3).value
                
                # 检查是否是单板开始行
                is_board_start = False
                board_model = None
                
                if c_val:
                    c_text = str(c_val)
                    # 检查是否包含单板型号信息
                    if "单板型号" in c_text or "型号：" in c_text:
                        is_board_start = True
                        board_model = self.extract_board_model(c_val)
                
                # 如果是单板开始行
                if is_board_start and board_model and board_model != '单板型号':
                    # 保存之前的单板
                    if current_board:
                        boards.append(current_board)
                    
                    # 创建新的单板
                    current_board = {
                        '序号': a_val if a_val else i,
                        '单板型号': board_model,
                        '接口信息': [],
                        'start_row': i,
                        'can_info': None  # 存储CAN接口信息
                    }
                    print(f"找到单板: 序号={a_val if a_val else i}, 型号={board_model}, 开始行={i}")
                    board_start_row = i
                    
                # 如果有当前单板，收集接口信息
                if current_board and board_start_row:
                    # 检查是否是接口行（E列有值且不是标题）
                    e_val = ws.cell(row=i, column=5).value
                    if (e_val and str(e_val).strip() != '' and 
                        str(e_val).strip() != 'SOC端' and 
                        str(e_val).strip() != 'MCU端' and
                        str(e_val).strip() != 'MCU端\n（CE架构时，接双芯片点位）'):
                        
                        # 获取完整的接口信息
                        interface_info = {
                            '行号': i,
                            'E列_PIN位': e_val,
                            'F列_连接器端': ws.cell(row=i, column=6).value,
                            'J列_信号名称': ws.cell(row=i, column=10).value,
                            'K列_功能说明': ws.cell(row=i, column=11).value,
                            'L列_功能说明': ws.cell(row=i, column=12).value,
                            'M列_功能说明': ws.cell(row=i, column=13).value,
                            'N列_功能说明': ws.cell(row=i, column=14).value,
                            'O列_功能说明': ws.cell(row=i, column=15).value,
                            'P列_接入类型': ws.cell(row=i, column=16).value
                        }
                        current_board['接口信息'].append(interface_info)
                    
                    # 检查CAN接口信息（在单板开始行后的几行中）
                    if i == board_start_row + 1:  # 通常在开始行后1行
                        p_val = ws.cell(row=i, column=16).value
                        if p_val and "接入" in str(p_val):
                            current_board['can_info'] = p_val
                            print(f"  找到CAN接口信息: {repr(p_val)}")
                        
            except Exception as ex:
                print(f"处理行 {i} 时出错: {ex}")
                continue
        
        # 添加最后一个单板
        if current_board:
            boards.append(current_board)
            
        wb.close()
        return boards
    
    def extract_pin_info(self, pin_str):
        """从PIN字符串中提取group和pin信息"""
        if not pin_str:
            return None, None
            
        # 处理格式如 "PB14" 的情况
        # 根据规则：PB14 -> group="GPB", pin="14"
        pin_str = str(pin_str).strip()
        
        # 处理带括号的格式，如 "CSI0_DAT19(L6)"
        bracket_match = re.search(r'\(([^)]+)\)$', pin_str)
        if bracket_match:
            # 如果有括号，提取括号内的内容作为PIN
            pin_part = bracket_match.group(1)
            # 提取字母和数字
            match = re.match(r'^([A-Z]+)(\d+)$', pin_part)
        else:
            # 直接处理格式如 "PB14"
            match = re.match(r'^([A-Z]+)(\d+)$', pin_str)
        
        if match:
            group = match.group(1)
            pin = match.group(2)
            # 转换组名：如PB -> GPB
            if group.startswith('P'):
                group = 'G' + group
            return group, pin
        return None, None
    
    def determine_interface_type(self, signal_name, klm_content):
        """根据信号名称和KL列内容确定接口类型"""
        if not signal_name:
            return "unknown"
            
        signal_name = str(signal_name).lower()
        klm_content = str(klm_content).lower() if klm_content else ""
        
        # 如果包含"IN"，则为DI类型
        if "in" in signal_name or "input" in signal_name:
            # 但是ADC-IN应该为AI类型
            if "adc-in" in signal_name or "analog" in signal_name:
                return "ai"
            return "di"
        
        # 如果包含"OUT"，则为DO类型
        if "out" in signal_name or "output" in signal_name:
            return "do"
            
        # 如果包含"adc"或"analog"，则为AI类型
        if "adc" in signal_name or "analog" in signal_name:
            return "ai"
            
        # 如果包含"can"，则为CAN通信接口
        if "can" in signal_name:
            return "can"
            
        # 根据KL列内容判断
        if "in" in klm_content:
            # 但是ADC-IN应该为AI类型
            if "adc" in klm_content or "analog" in klm_content:
                return "ai"
            return "di"
        elif "out" in klm_content:
            return "do"
            
        return "unknown"
    
    def process_single_board(self, board_info):
        """处理单个单板的配置"""
        print(f"\n处理单板: {board_info['单板型号']}")
        
        if not self.json_template:
            print("没有JSON模板数据")
            return None
            
        # 创建新的配置文件结构（使用模板）
        config = {
            "基本信息": {
                "name": board_info['单板型号'],
                "desc": "",
                "type": board_info['单板型号'],
                "ver": "V1.0.0",
                "id_dip": "ID_DIP_NULL",
                "ce_mode_set": 0,
                "di_hd_mode_set": 0,
                "board_type": [
                    "BOARD_SAFE_CTRL_MASTER_H8"
                ],
                "mcu": [
                    {
                        "chip": "R110"
                    }
                ]
            },
            "通信接口": {
                "can": []
            },
            "io接口": {
                "di": [],
                "do": [],
                "ai": []
            },
            "功能接口": {
                "pztb": [],
                "encr": []
            }
        }
        
        # 处理接口信息
        di_count = 0
        do_count = 0
        ai_count = 0
        
        for interface in board_info['接口信息']:
            e_pin = interface.get('E列_PIN位')
            j_signal = interface.get('J列_信号名称')
            p_access = interface.get('P列_接入类型')
            
            # 提取PIN信息
            group, pin = self.extract_pin_info(e_pin)
            
            # 确定接口类型
            interface_type = self.determine_interface_type(j_signal, interface.get('K列_功能说明'))
            
            print(f"  接口: {j_signal}, PIN: {e_pin}, Group: {group}, Pin: {pin}, Type: {interface_type}")
            
            # 根据接口类型添加到相应部分
            if interface_type == "di" and group and pin:
                di_count += 1
                di_entry = {
                    "name": f"DI_{di_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 0,
                    "freq": 0,
                    "pull_mode": 0,
                    "open_circle_level": 1,
                    "hd_if_mode": ""
                }
                config['io接口']['di'].append(di_entry)
                
            elif interface_type == "do" and group and pin:
                do_count += 1
                do_entry = {
                    "name": f"DO_{do_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 1,
                    "freq": 0,
                    "pull_mode": 0,
                    "enable_level": 1
                }
                config['io接口']['do'].append(do_entry)
                
            elif interface_type == "ai" and group and pin:
                ai_count += 1
                ai_entry = {
                    "name": f"AI_{ai_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 0,
                    "freq": 0,
                    "pull_mode": 0,
                    "adc_mode": 0,
                    "adc_ch": 0
                }
                config['io接口']['ai'].append(ai_entry)
                
            elif interface_type == "can" and p_access:
                # 处理CAN通信接口
                if "ucan" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_UCAN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
                elif "canopen" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_CANOPEN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
        
        # 处理单板级别的CAN信息
        if board_info.get('can_info'):
            can_access = board_info['can_info']
            print(f"  单板CAN信息: {can_access}")
            if "ucan" in str(can_access).lower():
                can_entry = {
                    "name": "CAN_1",
                    "protocol": "PROTOCOL_UCAN",
                    "ucan_id": [
                        {
                            "mcu": 0,
                            "peripheral": "hcan1",
                            "id_map": [66]
                        }
                    ]
                }
                config['通信接口']['can'].append(can_entry)
            elif "canopen" in str(can_access).lower():
                can_entry = {
                    "name": "CAN_1",
                    "protocol": "PROTOCOL_CANOPEN",
                    "ucan_id": [
                        {
                            "mcu": 0,
                            "peripheral": "hcan1",
                            "id_map": [66]
                        }
                    ]
                }
                config['通信接口']['can'].append(can_entry)
        
        return config
    
    def generate_config_files(self):
        """生成配置文件"""
        print("开始生成配置文件...")
        
        # 分析Excel结构
        boards = self.analyze_excel_for_boards()
        print(f"总共找到 {len(boards)} 个单板")
        
        # 为每个单板生成配置文件
        processed_count = 0
        for board in boards:
            if board['序号'] != 0:  # 只处理序号不为0的单板
                config = self.process_single_board(board)
                if config:
                    # 生成文件名（使用单板型号，替换特殊字符）
                    filename = f"{board['单板型号']}.json"
                    # 清理文件名中的非法字符
                    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                    filepath = os.path.join('output', filename)
                    
                    # 创建output目录
                    os.makedirs('output', exist_ok=True)
                    
                    # 保存配置文件
                    with open(filepath, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=2, ensure_ascii=False)
                    
                    print(f"已生成配置文件: {filename}")
                    processed_count += 1
        
        print(f"配置文件生成完成! 共处理了 {processed_count} 个单板")

def main():
    """主函数"""
    processor = OptimizedRobotConfigProcessor()
    processor.generate_config_files()

if __name__ == "__main__":
    main()
