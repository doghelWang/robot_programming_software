import openpyxl
import re

def analyze_detailed():
    """详细分析接口信息"""
    print("详细分析接口信息...")
    
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    # 查找包含"接入"的行
    print("查找包含'接入'的行:")
    for i in range(1, 200):
        p_val = ws.cell(row=i, column=16).value  # P列是接入类型
        if p_val and "接入" in str(p_val):
            c_val = ws.cell(row=i, column=3).value  # C列是单板型号
            j_val = ws.cell(row=i, column=10).value  # J列是信号名称
            print(f"行{i}: C={repr(c_val)}, J={repr(j_val)}, P={repr(p_val)}")
    
    # 特别查看202904335单板的P列
    print("\n详细查看202904335单板的P列:")
    for i in range(15, 50):  # 从单板开始行到后面几行
        c_val = ws.cell(row=i, column=3).value
        p_val = ws.cell(row=i, column=16).value
        j_val = ws.cell(row=i, column=10).value
        if p_val:
            print(f"行{i}: C={repr(c_val)}, J={repr(j_val)}, P={repr(p_val)}")
    
    wb.close()

if __name__ == "__main__":
    analyze_detailed()
