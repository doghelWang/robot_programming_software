import openpyxl

def debug_excel_structure():
    """调试Excel文件结构"""
    print("开始调试Excel文件结构...")
    
    # 加载工作簿
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    print("检查前50行的C列内容:")
    for i in range(1, 51):
        try:
            # 获取A、C、E、J列的值
            a_val = ws.cell(row=i, column=1).value
            c_val = ws.cell(row=i, column=3).value
            e_val = ws.cell(row=i, column=5).value
            j_val = ws.cell(row=i, column=10).value
            
            if c_val:
                print(f"行{i}: A={a_val}, C={repr(c_val)}, E={e_val}, J={j_val}")
                
        except Exception as ex:
            print(f"行{i}: 错误 - {ex}")
    
    wb.close()
    
    print("\n检查特定模式的行:")
    # 查找包含"单板型号"的行
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    for i in range(1, 100):
        try:
            c_val = ws.cell(row=i, column=3).value
            if c_val and "单板型号" in str(c_val):
                print(f"找到单板型号行 {i}: {repr(c_val)}")
                # 显示前后几行
                for j in range(max(1, i-2), min(100, i+5)):
                    val = ws.cell(row=j, column=3).value
                    if val:
                        print(f"  行{j}: {repr(val)}")
                print("---")
        except Exception as ex:
            continue
    
    wb.close()

if __name__ == "__main__":
    debug_excel_structure()
