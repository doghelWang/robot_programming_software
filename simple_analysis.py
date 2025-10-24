import openpyxl
import json

def analyze_excel_structure():
    """简单分析Excel结构来找出问题"""
    print("=== 分析Excel结构 ===")
    
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    print("检查关键行:")
    for i in range(10, 20):
        c_val = ws.cell(row=i, column=3).value
        p_val = ws.cell(row=i, column=16).value
        if c_val or p_val:
            print(f"行{i}: C={repr(c_val)}, P={repr(p_val)}")
    
    print("\n查找所有单板开始行:")
    for i in range(1, 50):
        c_val = ws.cell(row=i, column=3).value
        if c_val and ("单板型号" in str(c_val) or "型号：" in str(c_val)):
            print(f"单板开始行 {i}: {repr(c_val)}")
    
    wb.close()
    
    print("\n=== 分析完成 ===")

def check_files():
    """检查生成的文件"""
    print("=== 检查生成的文件 ===")
    
    # 检查202904335.json
    try:
        with open('output/202904335.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            print("202904335.json - CAN接口:")
            print(json.dumps(data['通信接口'], indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"读取202904335.json出错: {e}")
    
    # 检查202903931.json
    try:
        with open('output/202903931.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            print("\n202903931.json - CAN接口:")
            print(json.dumps(data['通信接口'], indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"读取202903931.json出错: {e}")

if __name__ == "__main__":
    analyze_excel_structure()
    print()
    check_files()
