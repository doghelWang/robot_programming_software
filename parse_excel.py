import openpyxl
import json
import re

def analyze_excel_structure():
    """分析Excel文件结构"""
    print("开始分析Excel文件结构...")
    
    # 加载工作簿
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    print("Excel工作表结构分析:")
    print("=" * 80)
    
    # 查看前50行的关键列
    print("行号 | A列(序号) | E列 | F列 | J列 | K列 | L列 | M列 | N列 | O列 | P列")
    print("-" * 80)
    
    for i in range(1, 51):
        try:
            a_val = ws.cell(row=i, column=1).value
            e_val = ws.cell(row=i, column=5).value
            f_val = ws.cell(row=i, column=6).value
            j_val = ws.cell(row=i, column=10).value
            k_val = ws.cell(row=i, column=11).value
            l_val = ws.cell(row=i, column=12).value
            m_val = ws.cell(row=i, column=13).value
            n_val = ws.cell(row=i, column=14).value
            o_val = ws.cell(row=i, column=15).value
            p_val = ws.cell(row=i, column=16).value
            
            # 处理None值
            a_str = str(a_val) if a_val is not None else "None"
            e_str = str(e_val) if e_val is not None else "None"
            f_str = str(f_val) if f_val is not None else "None"
            j_str = str(j_val) if j_val is not None else "None"
            k_str = str(k_val) if k_val is not None else "None"
            l_str = str(l_val) if l_val is not None else "None"
            m_str = str(m_val) if m_val is not None else "None"
            n_str = str(n_val) if n_val is not None else "None"
            o_str = str(o_val) if o_val is not None else "None"
            p_str = str(p_val) if p_val is not None else "None"
            
            print(f"{i:3d} | {a_str:8} | {e_str:10} | {f_str:10} | {j_str:10} | {k_str:10} | {l_str:10} | {m_str:10} | {n_str:10} | {o_str:10} | {p_str}")
            
        except Exception as ex:
            print(f"{i:3d} | Error - {ex}")
    
    wb.close()

def parse_excel_data():
    """解析Excel数据并提取所需信息"""
    print("\n开始解析Excel数据...")
    
    # 加载工作簿
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    # 存储结果
    results = []
    
    # 遍历所有行，寻找有效数据
    for i in range(1, 200):  # 假设最多200行
        try:
            # 获取A列（序号）和C列（单板型号）
            a_val = ws.cell(row=i, column=1).value
            c_val = ws.cell(row=i, column=3).value
            
            # 如果A列不为0或空，且C列有值，则创建文件
            if a_val is not None and a_val != 0 and a_val != '' and c_val is not None and c_val != '':
                print(f"找到有效行: 行{i}, A列={a_val}, C列={c_val}")
                
                # 提取该行及后续相关行的信息
                board_info = {
                    '序号': a_val,
                    '单板型号': c_val,
                    '接口信息': []
                }
                
                # 从当前行开始，查找该单板的所有接口信息
                current_row = i
                while current_row < 200:
                    # 检查是否还有该单板的接口信息
                    # 这里需要更复杂的逻辑来识别接口行
                    break  # 简化处理，先跳过详细解析
                
                results.append(board_info)
                
        except Exception as ex:
            # 忽略错误行，继续处理
            continue
    
    wb.close()
    return results

def analyze_json_structure():
    """分析JSON文件结构"""
    print("\n开始分析JSON文件结构...")
    
    with open('d:/received/RA-IC_I-A-1A3BH0.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print("JSON文件结构分析:")
    print("=" * 80)
    print(json.dumps(data, indent=2, ensure_ascii=False))
    
    return data

if __name__ == "__main__":
    # 分析Excel结构
    analyze_excel_structure()
    
    # 分析JSON结构
    json_data = analyze_json_structure()
    
    # 解析Excel数据
    excel_results = parse_excel_data()
    print(f"\n解析结果: 找到 {len(excel_results)} 个单板")
