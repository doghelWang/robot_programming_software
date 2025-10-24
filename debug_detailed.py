import openpyxl
import re

def debug_detailed():
    """详细调试Excel解析"""
    print("详细调试Excel解析...")
    
    # 加载工作簿
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    print("检查前20行的详细信息:")
    for i in range(1, 21):
        try:
            a_val = ws.cell(row=i, column=1).value
            c_val = ws.cell(row=i, column=3).value
            e_val = ws.cell(row=i, column=5).value
            j_val = ws.cell(row=i, column=10).value
            
            print(f"行{i}:")
            print(f"  A列(A序号): {repr(a_val)}")
            print(f"  C列(单板型号): {repr(c_val)}")
            print(f"  E列(PIN位): {repr(e_val)}")
            print(f"  J列(信号名称): {repr(j_val)}")
            
            if c_val:
                # 尝试提取单板型号
                c_text = str(c_val)
                print(f"  C列文本分析:")
                print(f"    原始文本: {repr(c_text)}")
                
                model_match = re.search(r'单板型号：([^\n]+)', c_text)
                if model_match:
                    model = model_match.group(1).strip()
                    print(f"    提取的单板型号: {repr(model)}")
                else:
                    print("    未找到'单板型号：'格式")
                    
                model_match2 = re.search(r'型号：([^\n]+)', c_text)
                if model_match2:
                    model2 = model_match2.group(1).strip()
                    print(f"    提取的型号: {repr(model2)}")
                else:
                    print("    未找到'型号：'格式")
                    
                lines = c_text.split('\n')
                print(f"    分行结果: {lines}")
                
            print()
            
        except Exception as ex:
            print(f"行{i}: 错误 - {ex}")
    
    wb.close()

if __name__ == "__main__":
    debug_detailed()
