import openpyxl
import re

def analyze_can_adc():
    """分析CAN和ADC接口"""
    print("分析CAN和ADC接口...")
    
    wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
    ws = wb['整机配置表']
    
    # 查找特定单板的CAN接口位置
    print("查找202904335单板的CAN接口信息:")
    for i in range(1, 200):
        c_val = ws.cell(row=i, column=3).value
        if c_val and "202904335" in str(c_val):
            print(f"找到202904335单板在行 {i}")
            # 查看该单板的P列信息
            for j in range(i, min(i+10, 200)):
                p_val = ws.cell(row=j, column=16).value
                if p_val:
                    print(f"  行{j} P列: {repr(p_val)}")
            break
    
    print("\n查找ADC-IN接口:")
    # 查找包含ADC-IN的接口
    for i in range(1, 200):
        j_val = ws.cell(row=i, column=10).value  # J列是信号名称
        p_val = ws.cell(row=i, column=16).value  # P列是接入类型
        if j_val and "ADC-IN" in str(j_val):
            print(f"找到ADC-IN接口 - 行{i}: J列={repr(j_val)}, P列={repr(p_val)}")
    
    wb.close()

if __name__ == "__main__":
    analyze_can_adc()
