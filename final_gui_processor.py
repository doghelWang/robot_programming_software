import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import json
import os
import re
from tkinter.scrolledtext import ScrolledText

class ExcelProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel配置文件处理器")
        self.root.geometry("800x600")
        
        # 数据存储
        self.excel_file_path = tk.StringVar()
        self.sheet_name = tk.StringVar(value="整机配置表")
        self.output_dir = tk.StringVar()
        self.json_template_path = tk.StringVar(value="d:/received/RA-IC_I-A-1A3BH0.json")
        
        self.setup_ui()
        
    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # Excel文件选择
        ttk.Label(file_frame, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.excel_file_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=5)
        ttk.Button(file_frame, text="浏览...", command=self.browse_excel_file).grid(row=0, column=2, sticky=tk.W, pady=5)
        
        # 工作表名称
        ttk.Label(file_frame, text="工作表名称:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.sheet_name, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=5)
        
        # JSON模板文件
        ttk.Label(file_frame, text="JSON模板:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.json_template_path, width=50).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=5)
        ttk.Button(file_frame, text="浏览...", command=self.browse_json_file).grid(row=2, column=2, sticky=tk.W, pady=5)
        
        # 输出目录
        ttk.Label(file_frame, text="输出目录:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_dir, width=50).grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=5)
        ttk.Button(file_frame, text="浏览...", command=self.browse_output_dir).grid(row=3, column=2, sticky=tk.W, pady=5)
        
        # 处理按钮
        process_frame = ttk.Frame(main_frame)
        process_frame.grid(row=1, column=0, columnspan=2, pady=10)
        
        self.process_button = ttk.Button(process_frame, text="开始处理", command=self.process_excel)
        self.process_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.cancel_button = ttk.Button(process_frame, text="取消", command=self.root.quit)
        self.cancel_button.pack(side=tk.LEFT)
        
        # 日志显示区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="10")
        log_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = ScrolledText(log_frame, height=20, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置滚动条
        self.log_text.config(state=tk.DISABLED)
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
    def log_message(self, message):
        """记录日志消息"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def browse_excel_file(self):
        """浏览Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_path.set(file_path)
            self.log_message(f"已选择Excel文件: {file_path}")
            
    def browse_json_file(self):
        """浏览JSON模板文件"""
        file_path = filedialog.askopenfilename(
            title="选择JSON模板文件",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if file_path:
            self.json_template_path.set(file_path)
            self.log_message(f"已选择JSON模板文件: {file_path}")
            
    def browse_output_dir(self):
        """浏览输出目录"""
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.output_dir.set(dir_path)
            self.log_message(f"已选择输出目录: {dir_path}")
            
    def load_json_template(self):
        """加载JSON模板"""
        try:
            with open(self.json_template_path.get(), 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            self.log_message(f"加载JSON模板失败: {e}")
            return None
            
    def extract_board_model(self, c_column_value):
        """从C列内容中提取单板型号"""
        if not c_column_value:
            return None
            
        c_text = str(c_column_value)
        
        # 查找"单板型号："格式
        model_match = re.search(r'单板型号：([^\n]+)', c_text)
        if model_match:
            return model_match.group(1).strip()
        
        # 查找"型号："格式
        model_match2 = re.search(r'型号：([^\n]+)', c_text)
        if model_match2:
            return model_match2.group(1).strip()
            
        # 如果没有找到特定格式，返回整个文本的前几部分
        lines = c_text.split('\n')
        if lines:
            return lines[0].strip()
            
        return None
    
    def analyze_excel_for_boards(self, wb, sheet_name):
        """分析Excel文件，识别所有单板"""
        self.log_message("分析Excel文件结构...")
        
        ws = wb[sheet_name]
        
        boards = []
        current_board = None
        board_start_row = None
        
        # 查找所有单板开始行（C列包含"单板型号"或"型号："的行）
        for i in range(1, 200):
            try:
                a_val = ws.cell(row=i, column=1).value
                c_val = ws.cell(row=i, column=3).value
                
                # 检查是否是单板开始行
                is_board_start = False
                board_model = None
                
                if c_val:
                    c_text = str(c_val)
                    # 检查是否包含单板型号信息
                    if "单板型号" in c_text or "型号：" in c_text:
                        is_board_start = True
                        board_model = self.extract_board_model(c_val)
                
                # 如果是单板开始行
                if is_board_start and board_model and board_model != '单板型号':
                    # 保存之前的单板
                    if current_board:
                        boards.append(current_board)
                    
                    # 创建新的单板
                    current_board = {
                        '序号': a_val if a_val else i,
                        '单板型号': board_model,
                        '接口信息': [],
                        'start_row': i,
                        'can_info': None  # 存储CAN接口信息
                    }
                    self.log_message(f"找到单板: 序号={a_val if a_val else i}, 型号={board_model}, 开始行={i}")
                    board_start_row = i
                    
                # 如果有当前单板，收集接口信息
                if current_board and board_start_row:
                    # 检查是否是接口行（E列有值且不是标题）
                    e_val = ws.cell(row=i, column=5).value
                    if (e_val and str(e_val).strip() != '' and 
                        str(e_val).strip() != 'SOC端' and 
                        str(e_val).strip() != 'MCU端' and
                        str(e_val).strip() != 'MCU端\n（CE架构时，接双芯片点位）'):
                        
                        # 获取完整的接口信息
                        interface_info = {
                            '行号': i,
                            'E列_PIN位': e_val,
                            'F列_连接器端': ws.cell(row=i, column=6).value,
                            'J列_信号名称': ws.cell(row=i, column=10).value,
                            'K列_功能说明': ws.cell(row=i, column=11).value,
                            'L列_功能说明': ws.cell(row=i, column=12).value,
                            'M列_功能说明': ws.cell(row=i, column=13).value,
                            'N列_功能说明': ws.cell(row=i, column=14).value,
                            'O列_功能说明': ws.cell(row=i, column=15).value,
                            'P列_接入类型': ws.cell(row=i, column=16).value
                        }
                        current_board['接口信息'].append(interface_info)
                    
                    # 检查CAN接口信息（根据新规则：从开始行向前3行开始查找）
                    # 例如：开始行是15，实际查找范围是12-15行
                    if board_start_row:
                        # 计算CAN查找的起始行（向前3行）
                        can_start_row = max(1, board_start_row - 3)
                        can_end_row = board_start_row + 30  # 保持原来的结束范围
                        
                        if can_start_row <= i <= can_end_row:
                            p_val = ws.cell(row=i, column=16).value
                            if p_val and "接入" in str(p_val):
                                current_board['can_info'] = p_val
                                self.log_message(f"  找到CAN接口信息: {repr(p_val)} (行{i})")
                        
            except Exception as ex:
                self.log_message(f"处理行 {i} 时出错: {ex}")
                continue
        
        # 添加最后一个单板
        if current_board:
            boards.append(current_board)
            
        return boards
    
    def extract_pin_info(self, pin_str):
        """从PIN字符串中提取group和pin信息"""
        if not pin_str:
            return None, None
            
        # 处理格式如 "PB14" 的情况
        # 根据规则：PB14 -> group="GPB", pin="14"
        pin_str = str(pin_str).strip()
        
        # 处理带括号的格式，如 "CSI0_DAT19(L6)"
        bracket_match = re.search(r'\(([^)]+)\)$', pin_str)
        if bracket_match:
            # 如果有括号，提取括号内的内容作为PIN
            pin_part = bracket_match.group(1)
            # 提取字母和数字
            match = re.match(r'^([A-Z]+)(\d+)$', pin_part)
        else:
            # 直接处理格式如 "PB14"
            match = re.match(r'^([A-Z]+)(\d+)$', pin_str)
        
        if match:
            group = match.group(1)
            pin = match.group(2)
            # 转换组名：如PB -> GPB
            if group.startswith('P'):
                group = 'G' + group
            return group, pin
        return None, None
    
    def determine_interface_type(self, signal_name, klm_content):
        """根据信号名称和KL列内容确定接口类型"""
        if not signal_name:
            return "unknown"
            
        signal_name = str(signal_name).lower()
        klm_content = str(klm_content).lower() if klm_content else ""
        
        # 如果包含"IN"，则为DI类型
        if "in" in signal_name or "input" in signal_name:
            # 但是ADC-IN应该为AI类型
            if "adc-in" in signal_name or "analog" in signal_name:
                return "ai"
            return "di"
        
        # 如果包含"OUT"，则为DO类型
        if "out" in signal_name or "output" in signal_name:
            return "do"
            
        # 如果包含"adc"或"analog"，则为AI类型
        if "adc" in signal_name or "analog" in signal_name:
            return "ai"
            
        # 如果包含"can"，则为CAN通信接口
        if "can" in signal_name:
            return "can"
            
        # 根据KL列内容判断
        if "in" in klm_content:
            # 但是ADC-IN应该为AI类型
            if "adc" in klm_content or "analog" in klm_content:
                return "ai"
            return "di"
        elif "out" in klm_content:
            return "do"
            
        return "unknown"
    
    def process_single_board(self, board_info, json_template):
        """处理单个单板的配置"""
        self.log_message(f"\n处理单板: {board_info['单板型号']}")
        
        if not json_template:
            self.log_message("没有JSON模板数据")
            return None
            
        # 创建新的配置文件结构（使用模板）
        config = {
            "基本信息": {
                "name": board_info['单板型号'],
                "desc": "",
                "type": board_info['单板型号'],
                "ver": "V1.0.0",
                "id_dip": "ID_DIP_NULL",
                "ce_mode_set": 0,
                "di_hd_mode_set": 0,
                "board_type": [
                    "BOARD_SAFE_CTRL_MASTER_H8"
                ],
                "mcu": [
                    {
                        "chip": "R110"
                    }
                ]
            },
            "通信接口": {
                "can": []
            },
            "io接口": {
                "di": [],
                "do": [],
                "ai": []
            },
            "功能接口": {
                "pztb": [],
                "encr": []
            }
        }
        
        # 处理接口信息
        di_count = 0
        do_count = 0
        ai_count = 0
        
        for interface in board_info['接口信息']:
            e_pin = interface.get('E列_PIN位')
            j_signal = interface.get('J列_信号名称')
            p_access = interface.get('P列_接入类型')
            
            # 提取PIN信息
            group, pin = self.extract_pin_info(e_pin)
            
            # 确定接口类型
            interface_type = self.determine_interface_type(j_signal, interface.get('K列_功能说明'))
            
            self.log_message(f"  接口: {j_signal}, PIN: {e_pin}, Group: {group}, Pin: {pin}, Type: {interface_type}")
            
            # 根据接口类型添加到相应部分
            if interface_type == "di" and group and pin:
                di_count += 1
                di_entry = {
                    "name": f"DI_{di_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 0,
                    "freq": 0,
                    "pull_mode": 0,
                    "open_circle_level": 1,
                    "hd_if_mode": ""
                }
                config['io接口']['di'].append(di_entry)
                
            elif interface_type == "do" and group and pin:
                do_count += 1
                do_entry = {
                    "name": f"DO_{do_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 1,
                    "freq": 0,
                    "pull_mode": 0,
                    "enable_level": 1
                }
                config['io接口']['do'].append(do_entry)
                
            elif interface_type == "ai" and group and pin:
                ai_count += 1
                ai_entry = {
                    "name": f"AI_{ai_count}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 0,
                    "freq": 0,
                    "pull_mode": 0,
                    "adc_mode": 0,
                    "adc_ch": 0
                }
                config['io接口']['ai'].append(ai_entry)
                
            elif interface_type == "can" and p_access:
                # 处理CAN通信接口
                if "ucan" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_UCAN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
                elif "canopen" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_CANOPEN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
        
        # 处理单板级别的CAN信息
        if board_info.get('can_info'):
            can_access = board_info['can_info']
            self.log_message(f"  单板CAN信息: {can_access}")
            if "ucan" in str(can_access).lower():
                can_entry = {
                    "name": "CAN_1",
                    "protocol": "PROTOCOL_UCAN",
                    "ucan_id": [
                        {
                            "mcu": 0,
                            "peripheral": "hcan1",
                            "id_map": [66]
                        }
                    ]
                }
                config['通信接口']['can'].append(can_entry)
            elif "canopen" in str(can_access).lower():
                can_entry = {
                    "name": "CAN_1",
                    "protocol": "PROTOCOL_CANOPEN",
                    "ucan_id": [
                        {
                            "mcu": 0,
                            "peripheral": "hcan1",
                            "id_map": [66]
                        }
                    ]
                }
                config['通信接口']['can'].append(can_entry)
        
        return config
    
    def generate_config_files(self, boards, json_template, output_dir):
        """生成配置文件"""
        self.log_message("开始生成配置文件...")
        
        # 为每个单板生成配置文件
        processed_count = 0
        for board in boards:
            if board['序号'] != 0:  # 只处理序号不为0的单板
                config = self.process_single_board(board, json_template)
                if config:
                    # 生成文件名（使用单板型号，替换特殊字符）
                    filename = f"{board['单板型号']}.json"
                    # 清理文件名中的非法字符
                    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                    filepath = os.path.join(output_dir, filename)
                    
                    # 保存配置文件
                    with open(filepath, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=2, ensure_ascii=False)
                    
                    self.log_message(f"已生成配置文件: {filename}")
                    processed_count += 1
        
        self.log_message(f"配置文件生成完成! 共处理了 {processed_count} 个单板")
        return processed_count
    
    def process_excel(self):
        """处理Excel文件"""
        # 检查输入
        if not self.excel_file_path.get():
            messagebox.showerror("错误", "请选择Excel文件")
            return
            
        if not self.output_dir.get():
            messagebox.showerror("错误", "请选择输出目录")
            return
            
        if not os.path.exists(self.excel_file_path.get()):
            messagebox.showerror("错误", "Excel文件不存在")
            return
            
        if not os.path.exists(self.output_dir.get()):
            messagebox.showerror("错误", "输出目录不存在")
            return
            
        # 禁用按钮
        self.process_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.DISABLED)
        self.status_var.set("处理中...")
        
        try:
            # 加载JSON模板
            json_template = self.load_json_template()
            
            # 加载Excel文件
            self.log_message(f"正在加载Excel文件: {self.excel_file_path.get()}")
            wb = openpyxl.load_workbook(self.excel_file_path.get(), read_only=True)
            
            # 分析Excel结构
            boards = self.analyze_excel_for_boards(wb, self.sheet_name.get())
            self.log_message(f"总共找到 {len(boards)} 个单板")
            
            # 生成配置文件
            processed_count = self.generate_config_files(boards, json_template, self.output_dir.get())
            
            # 完成
            wb.close()
            self.status_var.set(f"处理完成! 生成了 {processed_count} 个文件")
            messagebox.showinfo("完成", f"处理完成!\n生成了 {processed_count} 个配置文件")
            
        except Exception as e:
            self.log_message(f"处理过程中发生错误: {e}")
            messagebox.showerror("错误", f"处理过程中发生错误:\n{e}")
            self.status_var.set("错误")
        finally:
            # 恢复按钮状态
            self.process_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.NORMAL)

def main():
    root = tk.Tk()
    app = ExcelProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
