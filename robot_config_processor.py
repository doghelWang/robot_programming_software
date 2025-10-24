import openpyxl
import json
import os
import re

class RobotConfigProcessor:
    def __init__(self):
        self.json_data = None
        self.load_json_data()
    
    def load_json_data(self):
        """加载JSON数据"""
        try:
            with open('d:/received/RA-IC_I-A-1A3BH0.json', 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            print("JSON数据加载成功")
        except Exception as e:
            print(f"加载JSON数据失败: {e}")
            self.json_data = None
    
    def analyze_excel_structure(self):
        """分析Excel文件结构"""
        print("分析Excel文件结构...")
        
        # 加载工作簿
        wb = openpyxl.load_workbook('c:/Users/wangfeifei/Downloads/【323700510 MR-F0-50DCH-A7(M)】整机配置表-20240507.xlsx', read_only=True)
        ws = wb['整机配置表']
        
        # 查找所有单板信息
        boards = []
        current_board = None
        
        for i in range(1, 200):
            try:
                # 获取A列（序号）和C列（单板型号）
                a_val = ws.cell(row=i, column=1).value
                c_val = ws.cell(row=i, column=3).value
                
                # 检查是否是新的单板开始
                if a_val is not None and a_val != 0 and a_val != '' and c_val is not None and c_val != '':
                    if current_board:
                        boards.append(current_board)
                    
                    current_board = {
                        '序号': a_val,
                        '单板型号': c_val,
                        '接口信息': []
                    }
                    print(f"找到单板: 序号={a_val}, 型号={c_val}")
                    
                # 如果有当前单板，收集接口信息
                if current_board:
                    # 检查是否是接口行（E列有值）
                    e_val = ws.cell(row=i, column=5).value
                    f_val = ws.cell(row=i, column=6).value
                    j_val = ws.cell(row=i, column=10).value
                    k_val = ws.cell(row=i, column=11).value
                    l_val = ws.cell(row=i, column=12).value
                    m_val = ws.cell(row=i, column=13).value
                    n_val = ws.cell(row=i, column=14).value
                    o_val = ws.cell(row=i, column=15).value
                    p_val = ws.cell(row=i, column=16).value
                    
                    # 如果E列有值，说明是接口行
                    if e_val and str(e_val).strip() != '':
                        interface_info = {
                            '行号': i,
                            'E列_PIN位': e_val,
                            'F列_连接器端': f_val,
                            'J列_信号名称': j_val,
                            'K列_功能说明': k_val,
                            'L列_功能说明': l_val,
                            'M列_功能说明': m_val,
                            'N列_功能说明': n_val,
                            'O列_功能说明': o_val,
                            'P列_接入类型': p_val
                        }
                        current_board['接口信息'].append(interface_info)
                        print(f"  接口信息行 {i}: E={e_val}, J={j_val}")
                        
            except Exception as ex:
                continue
        
        if current_board:
            boards.append(current_board)
            
        wb.close()
        return boards
    
    def extract_pin_info(self, pin_str):
        """从PIN字符串中提取group和pin信息"""
        if not pin_str:
            return None, None
            
        # 处理格式如 "PB14" 的情况
        # 根据规则：PB14 -> group="GPB", pin="14"
        match = re.match(r'^([A-Z]+)(\d+)$', str(pin_str).strip())
        if match:
            group = match.group(1)
            pin = match.group(2)
            # 转换组名：如PB -> GPB
            if group.startswith('P'):
                group = 'G' + group
            return group, pin
        return None, None
    
    def determine_interface_type(self, signal_name, klm_content):
        """根据信号名称和KL列内容确定接口类型"""
        if not signal_name:
            return "unknown"
            
        signal_name = str(signal_name).lower()
        klm_content = str(klm_content).lower() if klm_content else ""
        
        # 如果包含"IN"，则为DI类型
        if "in" in signal_name or "input" in signal_name:
            return "di"
        
        # 如果包含"OUT"，则为DO类型
        if "out" in signal_name or "output" in signal_name:
            return "do"
            
        # 如果包含"adc"或"analog"，则为AI类型
        if "adc" in signal_name or "analog" in signal_name:
            return "ai"
            
        # 如果包含"can"，则为CAN通信接口
        if "can" in signal_name:
            return "can"
            
        # 根据KL列内容判断
        if "in" in klm_content:
            return "di"
        elif "out" in klm_content:
            return "do"
            
        return "unknown"
    
    def process_single_board(self, board_info):
        """处理单个单板的配置"""
        print(f"\n处理单板: {board_info['单板型号']}")
        
        # 从JSON模板中获取基础结构
        if not self.json_data:
            print("没有JSON模板数据")
            return None
            
        # 创建新的配置文件结构
        config = {
            "基本信息": {
                "name": board_info['单板型号'],
                "desc": "",
                "type": board_info['单板型号'],
                "ver": "V1.0.0",
                "id_dip": "ID_DIP_NULL",
                "ce_mode_set": 0,
                "di_hd_mode_set": 0,
                "board_type": [
                    "BOARD_SAFE_CTRL_MASTER_H8"
                ],
                "mcu": [
                    {
                        "chip": "R110"
                    }
                ]
            },
            "通信接口": {
                "can": []
            },
            "io接口": {
                "di": [],
                "do": [],
                "ai": []
            },
            "功能接口": {
                "pztb": [],
                "encr": []
            }
        }
        
        # 处理接口信息
        for interface in board_info['接口信息']:
            e_pin = interface.get('E列_PIN位')
            j_signal = interface.get('J列_信号名称')
            p_access = interface.get('P列_接入类型')
            
            # 提取PIN信息
            group, pin = self.extract_pin_info(e_pin)
            
            # 确定接口类型
            interface_type = self.determine_interface_type(j_signal, interface.get('K列_功能说明'))
            
            print(f"  接口: {j_signal}, PIN: {e_pin}, Group: {group}, Pin: {pin}, Type: {interface_type}")
            
            # 根据接口类型添加到相应部分
            if interface_type == "di" and group and pin:
                di_entry = {
                    "name": f"DI_{len(config['io接口']['di']) + 1}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 0,
                    "freq": 0,
                    "pull_mode": 0,
                    "open_circle_level": 1,
                    "hd_if_mode": ""
                }
                config['io接口']['di'].append(di_entry)
                
            elif interface_type == "do" and group and pin:
                do_entry = {
                    "name": f"DO_{len(config['io接口']['do']) + 1}",
                    "desc": j_signal or "",
                    "mcu": 0,
                    "group": group,
                    "pin": int(pin),
                    "io_mode": 1,
                    "freq": 0,
                    "pull_mode": 0,
                    "enable_level": 1
                }
                config['io接口']['do'].append(do_entry)
                
            elif interface_type == "can" and p_access:
                # 处理CAN通信接口
                if "ucan" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_UCAN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
                elif "canopen" in str(p_access).lower():
                    can_entry = {
                        "name": "CAN_1",
                        "protocol": "PROTOCOL_CANOPEN",
                        "ucan_id": [
                            {
                                "mcu": 0,
                                "peripheral": "hcan1",
                                "id_map": [66]
                            }
                        ]
                    }
                    config['通信接口']['can'].append(can_entry)
        
        return config
    
    def generate_config_files(self):
        """生成配置文件"""
        print("开始生成配置文件...")
        
        # 分析Excel结构
        boards = self.analyze_excel_structure()
        print(f"总共找到 {len(boards)} 个单板")
        
        # 为每个单板生成配置文件
        for board in boards:
            if board['序号'] != 0:  # 只处理序号不为0的单板
                config = self.process_single_board(board)
                if config:
                    # 生成文件名
                    filename = f"{board['单板型号']}.json"
                    filepath = os.path.join('output', filename)
                    
                    # 创建output目录
                    os.makedirs('output', exist_ok=True)
                    
                    # 保存配置文件
                    with open(filepath, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=2, ensure_ascii=False)
                    
                    print(f"已生成配置文件: {filename}")
        
        print("配置文件生成完成!")

def main():
    """主函数"""
    processor = RobotConfigProcessor()
    processor.generate_config_files()

if __name__ == "__main__":
    main()
